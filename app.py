from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from sqlalchemy import func, case
import io, hashlib, calendar
from datetime import date, datetime, timedelta
from functools import wraps
from waitress import serve
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from sheets_db import db, configure_app, pull_from_sheets, Employee, Attendance, Location, PO

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'attendance_secret_key_2024')
configure_app(app)

_db_initialized = False

def init_db():
    global _db_initialized
    if _db_initialized:
        return
    # Schema check/migration for SQLite
    with app.app_context():
        need_recreate = False
        try:
            # If locations doesn't have the zone column, we must recreate
            db.session.execute(db.text("SELECT zone FROM locations LIMIT 1"))
        except Exception:
            need_recreate = True
            
        try:
            # If pos table is missing or doesn't have reporting_person column, recreate
            db.session.execute(db.text("SELECT reporting_person FROM pos LIMIT 1"))
        except Exception:
            need_recreate = True
        
        try:
            db.session.execute(db.text("SELECT zone FROM attendance LIMIT 1"))
        except Exception:
            need_recreate = True

        try:
            db.session.execute(db.text("SELECT po_worked_days FROM attendance LIMIT 1"))
        except Exception:
            need_recreate = True

        try:
            db.session.execute(db.text("SELECT time_worked FROM attendance LIMIT 1"))
        except Exception:
            need_recreate = True

        try:
            db.session.execute(db.text("SELECT reporting_to FROM employees LIMIT 1"))
        except Exception:
            need_recreate = True

        try:
            db.session.execute(db.text("SELECT added_by FROM locations LIMIT 1"))
        except Exception:
            need_recreate = True

        try:
            db.session.execute(db.text("SELECT work_update FROM attendance LIMIT 1"))
        except Exception:
            need_recreate = True

        try:
            db.session.execute(db.text("SELECT added_by FROM pos LIMIT 1"))
        except Exception:
            need_recreate = True

        if need_recreate:
            print("Database schema mismatch detected. Dropping tables to recreate with new schema...")
            try:
                db.drop_all()
            except Exception as e:
                print(f"Error dropping tables: {e}")

        db.create_all()
        pull_from_sheets(app)
    
        if 'mssql' in str(db.engine.url):
            with db.engine.connect() as conn:
                conn.execute(db.text("""
                    IF NOT EXISTS (
                        SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                        WHERE TABLE_NAME = 'attendance' AND COLUMN_NAME = 'location'
                    )
                    BEGIN
                        ALTER TABLE attendance
                        ADD location VARCHAR(100) NULL
                    END
                """))
                conn.commit()
        # ──────────────────────────────────────────────────────────────────────────
        # Seed actual employees
        real_accounts = [
            ('2023002', 'Arun Venkatachalam', 'Management', 'Admin', 'arun.v@jdtech.co.in', 'admin'),
            ('2023017', 'Sanjay', 'SCADA', 'SCADA Engineer', 'sanjay.r@jdtech.co.in', 'employee'),
        ]
        for eid, name, dept, pos, email, role in real_accounts:
            pw = hashlib.sha256(eid.encode()).hexdigest()
            emp = Employee.query.filter_by(employee_id=eid).first()
            if not emp:
                db.session.add(Employee(
                    employee_id=eid,
                    password=pw,
                    role=role,
                    full_name=name,
                    department=dept,
                    position=pos,
                    email=email,
                    join_date=date(2023, 1, 1),
                    active=True
                ))
            else:
                if not emp.password:
                    emp.password = pw
                emp.role = role
                emp.active = True
                
        db.session.commit()
        _db_initialized = True

# Auto-initialize database on application import (for production WSGI servers like Gunicorn)
try:
    init_db()
except Exception as _e:
    print(f"Notice during database initialization: {_e}")

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Admin access required.', 'error')
            return redirect(url_for('my_attendance'))
        return f(*args, **kwargs)
    return decorated

def admin_or_tl_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') not in ['admin', 'tl']:
            flash('Admin or TL access required.', 'error')
            return redirect(url_for('my_attendance'))
        return f(*args, **kwargs)
    return decorated

def calculate_time_worked(check_in, check_out):
    if not check_in or not check_out:
        return None
    try:
        ci_h, ci_m = map(int, check_in.split(':'))
        co_h, co_m = map(int, check_out.split(':'))
        diff_mins = (co_h * 60 + co_m) - (ci_h * 60 + ci_m)
        if diff_mins < 0:
            diff_mins += 24 * 60
        h = diff_mins // 60
        m = diff_mins % 60
        return f"{h}h {m}m"
    except Exception:
        return None

def recalculate_employee_po_worked_days(employee_id):
    if not employee_id:
        return
    records = Attendance.query.filter_by(employee_id=employee_id).order_by(Attendance.date.asc()).all()
    po_counts = {}
    for r in records:
        if r.po_number:
            po = r.po_number.strip()
            if not po:
                r.po_worked_days = None
                continue
            if r.status in ['Present', 'Half Day']:
                po_counts[po] = po_counts.get(po, 0) + 1
                r.po_worked_days = po_counts[po]
            else:
                r.po_worked_days = None
        else:
            r.po_worked_days = None

@app.route('/')
def index():
    if 'user_id' in session:
        if session.get('role') == 'admin':
            return redirect(url_for('dashboard'))
        else:
            return redirect(url_for('my_attendance'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        login_input = request.form['username'].strip()
        pw = hashlib.sha256(request.form['password'].encode()).hexdigest()
        emp = Employee.query.filter(
            (Employee.employee_id == login_input) &
            (Employee.password == pw) &
            (Employee.active == True)
        ).first()
        if emp:
            session.update(
                user_id=emp.id,
                username=emp.employee_id,
                role=emp.role,
                full_name=emp.full_name,
                employee_id=emp.employee_id
            )
            if emp.role == 'admin':
                return redirect(url_for('my_attendance'))
            else:
                return redirect(url_for('my_attendance'))
        flash('Invalid Employee ID or password.', 'error')
    return render_template('login.html')

@app.route('/entry', methods=['GET', 'POST'])
@login_required
def attendance_entry():
    return redirect(url_for('my_attendance'))

    return render_template('attendance_entry.html',
        today=today.isoformat(), existing=existing,
        employee=employee, employee_id=employee_id)

@app.route('/my-attendance', methods=['GET', 'POST'])
@login_required
def my_attendance():
    try:
        employee_id = session.get('employee_id')
        employee    = None
        records     = []
        today       = date.today()
        today_record = None

        pending_work_update_record = None
        if employee_id:
            employee = Employee.query.filter_by(employee_id=employee_id).first()
            pending_work_update_record = Attendance.query.filter(
                Attendance.employee_id == employee_id,
                Attendance.date < today,
                Attendance.status.in_(['Present', 'Half Day']),
                (Attendance.work_update == None) | (func.trim(Attendance.work_update) == '')
            ).order_by(Attendance.date.desc()).first()

        if request.method == 'POST' and employee:
            action = request.form.get('action')
            existing = Attendance.query.filter_by(
                employee_id=employee_id, date=today).first()

            if action == 'clock_out':
                if existing:
                    if existing.check_out:
                        flash('Already clocked out today!', 'info')
                    else:
                        now_time = datetime.now().strftime('%H:%M')
                        existing.check_out = now_time
                        existing.time_worked = calculate_time_worked(existing.check_in, now_time)
                        
                        notes_from_form = request.form.get('notes', '').strip()
                        if notes_from_form:
                            if existing.notes:
                                existing.notes = f"{existing.notes} | {notes_from_form}"
                            else:
                                existing.notes = notes_from_form
                        
                        work_update_from_form = request.form.get('work_update', '').strip()
                        if work_update_from_form:
                            existing.work_update = work_update_from_form
                        
                        recalculate_employee_po_worked_days(employee_id)
                        db.session.commit()
                        flash(f'✅ Clock Out recorded successfully at {now_time}!', 'success')
                else:
                    flash('No check-in record found for today!', 'error')
                return redirect(url_for('my_attendance'))

            if existing:
                flash('⚠️ Attendance has already been logged for today! Only 1 attendance entry per person per day is permitted.', 'error')
                return redirect(url_for('my_attendance'))
            else:
                status           = request.form.get('status', 'Present')
                location         = request.form.get('location', '')
                po_number        = request.form.get('po_number', '').strip() or request.form.get('po_number_text', '').strip()
                reporting_person = request.form.get('reporting_person', '').strip()
                zone             = request.form.get('zone', '').strip() or request.form.get('zone_text', '').strip()
                notes            = request.form.get('notes', '')
                work_update      = request.form.get('work_update', '').strip()

                if pending_work_update_record and status in ['Present', 'Half Day']:
                    flash(f'⚠️ Work update for {pending_work_update_record.date.strftime("%d %b %Y")} is missing! Please submit your work update for that day before clocking in.', 'error')
                    return redirect(url_for('my_attendance'))

                if status == 'Absent':
                    location         = ''
                    po_number        = ''
                    reporting_person = ''
                    zone             = ''
                else:
                    # Server-side GPS verification
                    loc = Location.query.filter_by(customer_name=location).first()
                    if loc:
                        user_lat = request.form.get('user_lat')
                        user_lon = request.form.get('user_lon')
                        if not loc.latitude or not loc.longitude:
                            # Coordinates not set yet: register them from the user's manual entry/capture
                            if user_lat and user_lon:
                                try:
                                    loc.latitude = float(user_lat)
                                    loc.longitude = float(user_lon)
                                    if not loc.added_by:
                                        loc.added_by = session.get('full_name') or session.get('username') or session.get('employee_id')
                                    db.session.commit()
                                    flash(f'GPS coordinates registered for location "{location}". Location verification is now locked and active.', 'info')
                                except (TypeError, ValueError):
                                    pass
                        else:
                            # Standard GPS Verification
                            try:
                                lat = float(user_lat)
                                lon = float(user_lon)
                                dist = haversine_distance(lat, lon, loc.latitude, loc.longitude)
                                if dist > 200:
                                    flash(f'Out of location range! You are {round(dist)} m away from {location}.', 'error')
                                    return redirect(url_for('my_attendance'))
                            except (TypeError, ValueError):
                                flash('Could not verify GPS location. Please make sure location permissions are enabled.', 'error')
                                return redirect(url_for('my_attendance'))

                now_time = datetime.now().strftime('%H:%M')

                # Auto-add new PO to pos table if typed manually and not available
                if po_number:
                    existing_po = PO.query.filter_by(po_number=po_number).first()
                    if not existing_po:
                        max_sno = db.session.query(func.max(PO.s_no)).scalar() or 0
                        user_name = session.get('full_name') or session.get('username') or session.get('employee_id')
                        db.session.add(PO(
                            s_no=max_sno + 1,
                            customer_name=location or 'General',
                            zone=zone or '',
                            po_number=po_number,
                            days='',
                            reporting_person=reporting_person or '',
                            added_by=user_name
                        ))

                db.session.add(Attendance(
                    employee_id=employee_id,
                    date=today,
                    status=status,
                    check_in=now_time,
                    location=location,
                    po_number=po_number,
                    reporting_person=reporting_person,
                    zone=zone,
                    notes=notes,
                    work_update=work_update,
                    marked_by=session.get('username', 'Self')
                ))
                recalculate_employee_po_worked_days(employee_id)
                db.session.commit()
                flash('✅ Attendance manually recorded successfully!', 'success')
            return redirect(url_for('my_attendance'))

        if employee:
            records = Attendance.query.filter_by(employee_id=employee_id)\
                .order_by(Attendance.date.desc()).limit(60).all()
            today_record = Attendance.query.filter_by(
                employee_id=employee_id, date=today).first()

        present_count = sum(1 for r in records if r.status == 'Present')
        absent_count  = sum(1 for r in records if r.status == 'Absent')
        halfday_count = sum(1 for r in records if r.status == 'Half Day')

        raw_locations = Location.query.all()
        locations = []
        seen_locs = set()
        for l in raw_locations:
            if not l.customer_name:
                continue
            if l.customer_name not in seen_locs:
                seen_locs.add(l.customer_name)
                zones_list = [z.strip() for z in l.zone.split(',') if z.strip()] if l.zone else []
                locations.append({
                    'name': l.customer_name,
                    'customer_name': l.customer_name,
                    'zones': zones_list,
                    'gps_enabled': bool(l.latitude and l.longitude),
                    'latitude': l.latitude,
                    'longitude': l.longitude,
                    'radius_m': 200
                })
        locations.sort(key=lambda x: x['name'])

        pos = []
        raw_pos = PO.query.all()
        for p in raw_pos:
            pos.append({
                'customer_name': p.customer_name,
                'zone': p.zone,
                'po_number': p.po_number,
                'days': p.days,
                'reporting_person': p.reporting_person
            })
            
        reporting_persons = []
        location_checking_enabled = any(loc['gps_enabled'] for loc in locations)

        return render_template('my_attendance.html',
            employee=employee, records=records, today=today.isoformat(),
            today_record=today_record, present_count=present_count,
            absent_count=absent_count, halfday_count=halfday_count,
            locations=locations, pos=pos, reporting_persons=reporting_persons,
            location_checking_enabled=location_checking_enabled,
            pending_work_update_record=pending_work_update_record)
    except Exception as e:
        import traceback
        print("EXCEPTION IN MY_ATTENDANCE:")
        traceback.print_exc()
        raise e

@app.route('/update-work-update', methods=['POST'])
@login_required
def update_work_update():
    record_id = request.form.get('record_id')
    work_update_text = request.form.get('work_update', '').strip()
    employee_id = session.get('employee_id')
    
    if not record_id or not work_update_text:
        flash('Work update text cannot be empty.', 'error')
        return redirect(url_for('my_attendance'))
        
    record = Attendance.query.filter_by(id=record_id).first()
    if record:
        if session.get('role') != 'admin' and record.employee_id != employee_id:
            flash('Unauthorized to update this attendance record.', 'error')
            return redirect(url_for('my_attendance'))
        
        record.work_update = work_update_text
        db.session.commit()
        flash(f'✅ Work update saved for {record.date.strftime("%d %b %Y")}!', 'success')
    else:
        flash('Attendance record not found.', 'error')
        
    return redirect(url_for('my_attendance'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
@admin_or_tl_required
def dashboard():
    today         = date.today()
    role          = session.get('role')
    
    if role == 'tl':
        team_members = Employee.query.filter_by(reporting_to=session['employee_id'], active=True).all()
        team_ids = [e.employee_id for e in team_members]
        
        total_emp = len(team_members)
        if team_ids:
            present_today = Attendance.query.filter(Attendance.date == today, Attendance.status == 'Present', Attendance.employee_id.in_(team_ids)).count()
            absent_today  = Attendance.query.filter(Attendance.date == today, Attendance.status == 'Absent', Attendance.employee_id.in_(team_ids)).count()
        else:
            present_today = 0
            absent_today  = 0
    else:
        total_emp     = Employee.query.filter_by(active=True).count()
        present_today = Attendance.query.filter_by(date=today, status='Present').count()
        absent_today  = Attendance.query.filter_by(date=today, status='Absent').count()
        
    not_marked    = total_emp - present_today - absent_today

    trend = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        if role == 'tl':
            if team_ids:
                p = Attendance.query.filter(Attendance.date == d, Attendance.status == 'Present', Attendance.employee_id.in_(team_ids)).count()
            else:
                p = 0
        else:
            p = Attendance.query.filter_by(date=d, status='Present').count()
        trend.append({'date': d.isoformat(), 'present': p})

    recent_q_base = db.session.query(Attendance, Employee).join(Employee, Attendance.employee_id == Employee.employee_id)
    if role == 'tl':
        if team_ids:
            recent_q_base = recent_q_base.filter(Attendance.employee_id.in_(team_ids))
        else:
            recent_q_base = recent_q_base.filter(db.text("1=0"))
            
    recent_q = recent_q_base.order_by(Attendance.date.desc(), Attendance.id.desc()).limit(10).all()
    
    recent = [{
        'employee_id': a.employee_id, 'date': a.date.isoformat(), 'status': a.status,
        'location': a.location or '',
        'po_number': a.po_number or '',
        'po_worked_days': a.po_worked_days or '—',
        'time_worked': a.time_worked or '—',
        'reporting_person': a.reporting_person or '',
        'zone': a.zone or '',
        'full_name': e.full_name, 'department': e.department
    } for a, e in recent_q]

    return render_template('dashboard.html',
        total=total_emp, present=present_today, absent=absent_today,
        not_marked=not_marked, trend=trend, recent=recent, today=today.isoformat())

@app.route('/api/dashboard/stats')
@login_required
@admin_or_tl_required
def dashboard_stats():
    today         = date.today()
    role          = session.get('role')
    
    if role == 'tl':
        team_members = Employee.query.filter_by(reporting_to=session['employee_id'], active=True).all()
        team_ids = [e.employee_id for e in team_members]
        
        total_emp = len(team_members)
        if team_ids:
            present_today = Attendance.query.filter(Attendance.date == today, Attendance.status == 'Present', Attendance.employee_id.in_(team_ids)).count()
            absent_today  = Attendance.query.filter(Attendance.date == today, Attendance.status == 'Absent', Attendance.employee_id.in_(team_ids)).count()
        else:
            present_today = 0
            absent_today  = 0
    else:
        total_emp     = Employee.query.filter_by(active=True).count()
        present_today = Attendance.query.filter_by(date=today, status='Present').count()
        absent_today  = Attendance.query.filter_by(date=today, status='Absent').count()
        
    not_marked    = total_emp - present_today - absent_today

    recent_q_base = db.session.query(Attendance, Employee).join(Employee, Attendance.employee_id == Employee.employee_id)
    if role == 'tl':
        if team_ids:
            recent_q_base = recent_q_base.filter(Attendance.employee_id.in_(team_ids))
        else:
            recent_q_base = recent_q_base.filter(db.text("1=0"))
            
    recent_q = recent_q_base.order_by(Attendance.date.desc(), Attendance.id.desc()).limit(10).all()
    
    recent = [{
        'employee_id': a.employee_id, 'date': a.date.isoformat(), 'status': a.status,
        'location': a.location or '',
        'po_number': a.po_number or '',
        'po_worked_days': a.po_worked_days or '—',
        'time_worked': a.time_worked or '—',
        'reporting_person': a.reporting_person or '',
        'zone': a.zone or '',
        'full_name': e.full_name, 'department': e.department
    } for a, e in recent_q]

    return jsonify({
        'total': total_emp,
        'present': present_today,
        'absent': absent_today,
        'not_marked': not_marked,
        'recent': recent
    })

@app.route('/mark', methods=['GET', 'POST'])
@login_required
@admin_or_tl_required
def mark_attendance():
    today = date.today()
    role = session.get('role')
    
    if request.method == 'POST':
        sel_date  = date.fromisoformat(request.form.get('date', today.isoformat()))
        if role == 'tl':
            employees = Employee.query.filter_by(reporting_to=session['employee_id'], active=True).order_by(
                Employee.employee_id.asc()).all()
        else:
            employees = Employee.query.filter_by(active=True).order_by(
                Employee.employee_id.asc()).all()
                
        updated = 0
        modified_emp_ids = set()
        for emp in employees:
            status    = request.form.get(f'status_{emp.employee_id}')
            check_in  = request.form.get(f'checkin_{emp.employee_id}', '')
            check_out = request.form.get(f'checkout_{emp.employee_id}', '')
            notes     = request.form.get(f'notes_{emp.employee_id}', '')
            location  = request.form.get(f'location_{emp.employee_id}', '')
            if status == 'Absent':
                location = ''
                check_in = ''
                check_out = ''
            if status:
                po_obj = PO.query.filter_by(customer_name=location).first() if location else None
                loc_obj = Location.query.filter_by(customer_name=location).first() if location else None
                zone_val = po_obj.zone if po_obj else (loc_obj.zone if loc_obj else '')
                po_num_val = po_obj.po_number if po_obj else ''
                rp_val = po_obj.reporting_person if po_obj else ''

                rec = Attendance.query.filter_by(
                    employee_id=emp.employee_id, date=sel_date).first()
                if rec:
                    rec.status, rec.check_in, rec.check_out = status, check_in, check_out
                    rec.notes, rec.marked_by, rec.location = notes, session['username'], location
                    rec.zone = zone_val
                    rec.po_number = po_num_val
                    rec.reporting_person = rp_val
                    rec.time_worked = calculate_time_worked(check_in, check_out)
                else:
                    db.session.add(Attendance(
                        employee_id=emp.employee_id, date=sel_date, status=status,
                        check_in=check_in, check_out=check_out,
                        time_worked=calculate_time_worked(check_in, check_out),
                        notes=notes, marked_by=session['username'], location=location,
                        zone=zone_val, po_number=po_num_val, reporting_person=rp_val))
                updated += 1
                modified_emp_ids.add(emp.employee_id)
        
        for emp_id in modified_emp_ids:
            recalculate_employee_po_worked_days(emp_id)
        db.session.commit()
        flash(f'Attendance saved for {updated} employees on {sel_date}.', 'success')
        return redirect(url_for('mark_attendance'))

    selected_date = request.args.get('date', today.isoformat())
    sel_date_obj  = date.fromisoformat(selected_date)
    
    if role == 'tl':
        employees = Employee.query.filter_by(reporting_to=session['employee_id'], active=True).order_by(
            Employee.employee_id.asc()).all()
    else:
        employees = Employee.query.filter_by(active=True).order_by(
            Employee.employee_id.asc()).all()
            
    existing      = {r.employee_id: r for r in
                     Attendance.query.filter_by(date=sel_date_obj).all()}
    raw_locs = Location.query.all()
    seen_sites = set()
    sites = []
    for loc in raw_locs:
        if not loc.customer_name:
            continue
        display_name = loc.customer_name
        if display_name not in seen_sites:
            seen_sites.add(display_name)
            sites.append(display_name)
    sites.sort()
    return render_template('mark_attendance.html',
        employees=employees, existing=existing, selected_date=selected_date,
        sites=sites)

@app.route('/employees')
@login_required
@admin_or_tl_required
def employees():
    role = session.get('role')
    if role == 'tl':
        team = Employee.query.filter_by(reporting_to=session['employee_id'], active=True).order_by(
            Employee.employee_id.asc()).all()
        if team:
            emps = team
        else:
            emps = Employee.query.filter_by(active=True).order_by(
                Employee.employee_id.asc()).all()
    else:
        emps = Employee.query.filter_by(active=True).order_by(
            Employee.employee_id.asc()).all()
    return render_template('employees.html', employees=emps)

@app.route('/employees/add', methods=['GET', 'POST'])
@app.route('/employees/edit/<emp_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def add_employee(emp_id=None):
    target_emp = None
    if emp_id:
        target_emp = Employee.query.filter_by(employee_id=emp_id).first_or_404()

    if request.method == 'POST':
        try:
            input_emp_id = request.form['employee_id'].strip()
            full_name = request.form['full_name'].strip()
            dept = request.form['department'].strip()
            email = request.form.get('email', '').strip()
            phone = request.form.get('phone', '').strip()
            role = request.form.get('role', 'employee')
            pw = request.form.get('password')
            reporting_to = request.form.get('reporting_to', '').strip() or None
            position = request.form.get('position', '').strip() or None
            jd = date.fromisoformat(request.form['join_date']) if request.form.get('join_date') else None

            existing = target_emp or Employee.query.filter_by(employee_id=input_emp_id).first()
            if existing:
                existing.employee_id = input_emp_id
                existing.full_name = full_name
                existing.department = dept
                existing.position = position
                existing.email = email
                existing.phone = phone
                existing.role = role
                existing.reporting_to = reporting_to
                existing.active = True
                if pw and pw.strip():
                    existing.password = hashlib.sha256(pw.strip().encode()).hexdigest()
                if jd:
                    existing.join_date = jd
                db.session.commit()
                flash(f'Employee {full_name} ({input_emp_id}) updated successfully!', 'success')
            else:
                if not pw or not pw.strip():
                    pw = input_emp_id
                pw_hash = hashlib.sha256(pw.strip().encode()).hexdigest()
                db.session.add(Employee(
                    employee_id=input_emp_id,
                    password=pw_hash,
                    role=role,
                    full_name=full_name,
                    department=dept,
                    position=position,
                    email=email,
                    phone=phone,
                    join_date=jd,
                    active=True,
                    reporting_to=reporting_to))
                db.session.commit()
                flash(f'Employee {full_name} ({input_emp_id}) added successfully!', 'success')
            return redirect(url_for('employees'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'error')
            
    tls = Employee.query.filter_by(role='tl', active=True).order_by(Employee.employee_id.asc()).all()
    return render_template('add_employee.html', tls=tls, emp=target_emp)

@app.route('/employees/delete/<emp_id>', methods=['POST'])
@login_required
@admin_required
def delete_employee(emp_id):
    emp = Employee.query.filter_by(employee_id=emp_id).first_or_404()
    db.session.delete(emp)
    db.session.commit()
    flash('Employee removed from system and Google Sheets.', 'success')
    return redirect(url_for('employees'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_pw = request.form.get('current_password', '').strip()
        new_pw = request.form.get('new_password', '').strip()
        confirm_pw = request.form.get('confirm_password', '').strip()

        curr_hash = hashlib.sha256(current_pw.encode()).hexdigest()
        emp_id = session.get('employee_id')
        emp = Employee.query.filter_by(employee_id=emp_id).first()

        if not emp or emp.password != curr_hash:
            flash('Current password is incorrect.', 'error')
        elif not new_pw or len(new_pw) < 4:
            flash('New password must be at least 4 characters long.', 'error')
        elif new_pw != confirm_pw:
            flash('New passwords do not match.', 'error')
        else:
            emp.password = hashlib.sha256(new_pw.encode()).hexdigest()
            db.session.commit()
            flash('✅ Password updated successfully!', 'success')
            if session.get('role') == 'admin':
                return redirect(url_for('dashboard'))
            return redirect(url_for('my_attendance'))

    return render_template('change_password.html')

@app.route('/reports')
@login_required
@admin_or_tl_required
def reports():
    filter_type = request.args.get('filter_type', 'monthly')
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    dept    = request.args.get('dept', 'All')
    role = session.get('role')

    if filter_type == 'custom' and start_date_str and end_date_str:
        try:
            start = date.fromisoformat(start_date_str)
            end = date.fromisoformat(end_date_str)
        except ValueError:
            start = date.today().replace(day=1)
            end = date.today()
            start_date_str = start.isoformat()
            end_date_str = end.isoformat()
    else:
        try:
            y, m    = map(int, month.split('-'))
            start   = date(y, m, 1)
            end     = date(y, m, calendar.monthrange(y, m)[1])
        except Exception:
            start   = date.today().replace(day=1)
            end     = date.today()
            month   = date.today().strftime('%Y-%m')
        start_date_str = start.isoformat()
        end_date_str = end.isoformat()

    end_day = (end - start).days + 1

    query = (
        db.session.query(
            Employee.employee_id, Employee.full_name, Employee.department,
            func.sum(case((Attendance.status == 'Present',  1), else_=0)).label('present_days'),
            func.sum(case((Attendance.status == 'Absent',   1), else_=0)).label('absent_days'),
            func.sum(case((Attendance.status == 'Half Day', 1), else_=0)).label('halfday_days'),
            func.count(Attendance.id).label('total_marked'),
        )
        .outerjoin(Attendance,
            (Employee.employee_id == Attendance.employee_id) &
            (Attendance.date.between(start, end)))
        .filter(Employee.active == True)
    )
    if role == 'tl':
        query = query.filter(Employee.reporting_to == session['employee_id'])
    if dept != 'All':
        query = query.filter(Employee.department == dept)
    rows = query.group_by(
        Employee.employee_id, Employee.full_name, Employee.department
    ).order_by(Employee.employee_id.asc()).all()

    if role == 'tl':
        depts = [r[0] for r in
                 db.session.query(Employee.department).filter_by(reporting_to=session['employee_id'], active=True)
                 .distinct().order_by(Employee.department).all()]
    else:
        depts = [r[0] for r in
                 db.session.query(Employee.department).filter_by(active=True)
                 .distinct().order_by(Employee.department).all()]

    # Location breakdown for the period
    loc_query = (
        db.session.query(Attendance.location, func.count(Attendance.id).label('count'))
        .join(Employee, Attendance.employee_id == Employee.employee_id)
        .filter(Employee.active == True)
        .filter(Attendance.date.between(start, end))
        .filter(Attendance.location != None)
        .filter(Attendance.location != '')
    )
    if role == 'tl':
        loc_query = loc_query.filter(Employee.reporting_to == session['employee_id'])
    if dept != 'All':
        loc_query = loc_query.filter(Employee.department == dept)
    loc_rows = loc_query.group_by(Attendance.location).order_by(func.count(Attendance.id).desc()).all()

    # Detailed per-day records for the detailed table
    detail_query = (
        db.session.query(
            Employee.employee_id, Employee.full_name, Employee.department,
            Attendance.date, Attendance.status, Attendance.check_in, Attendance.check_out,
            Attendance.time_worked, Attendance.location, Attendance.po_number,
            Attendance.reporting_person, Attendance.zone, Attendance.notes
        )
        .join(Employee, Attendance.employee_id == Employee.employee_id)
        .filter(Employee.active == True)
        .filter(Attendance.date.between(start, end))
        .filter(Attendance.status.in_(['Present', 'Half Day']))
    )
    if role == 'tl':
        detail_query = detail_query.filter(Employee.reporting_to == session['employee_id'])
    if dept != 'All':
        detail_query = detail_query.filter(Employee.department == dept)
    detail_rows = detail_query.order_by(Attendance.date.desc(), Employee.full_name).all()

    return render_template('reports.html',
        rows=rows, month=month, start_date=start_date_str, end_date=end_date_str,
        filter_type=filter_type, dept=dept, depts=depts, end_day=end_day,
        loc_rows=loc_rows, detail_rows=detail_rows)

@app.route('/export')
@login_required
@admin_or_tl_required
def export_csv():
    filter_type = request.args.get('filter_type', 'monthly')
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    dept    = request.args.get('dept', 'All')
    export_format = request.args.get('format', 'pdf')
    role = session.get('role')

    if filter_type == 'custom' and start_date_str and end_date_str:
        try:
            start = date.fromisoformat(start_date_str)
            end = date.fromisoformat(end_date_str)
        except ValueError:
            start = date.today().replace(day=1)
            end = date.today()
    else:
        try:
            y, m    = map(int, month.split('-'))
            start   = date(y, m, 1)
            end     = date(y, m, calendar.monthrange(y, m)[1])
        except Exception:
            start   = date.today().replace(day=1)
            end     = date.today()

    end_day = (end - start).days + 1

    query = (
        db.session.query(
            Employee.employee_id, Employee.full_name, Employee.department,
            Attendance.date, Attendance.status, Attendance.check_in, Attendance.check_out,
            Attendance.time_worked, Attendance.location, Attendance.po_number,
            Attendance.po_worked_days, Attendance.reporting_person, Attendance.zone, Attendance.notes)
        .outerjoin(Attendance,
            (Employee.employee_id == Attendance.employee_id) &
            (Attendance.date.between(start, end)))
        .filter(Employee.active == True)
    )
    if role == 'tl':
        query = query.filter(Employee.reporting_to == session['employee_id'])
    if dept != 'All':
        query = query.filter(Employee.department == dept)
    rows = query.order_by(Employee.employee_id.asc(), Attendance.date).all()

    # ── Excel Export (xlsx) ───────────────────────────────────────────────────
    if export_format == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            ws.views.sheetView[0].showGridLines = True

            # Title Block
            ws.merge_cells('A1:L1')
            title_cell = ws['A1']
            title_cell.value = "JD TECH — Attendance Report"
            title_cell.font = Font(name="Segoe UI", size=15, bold=True, color="00D4AA")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells('A2:L2')
            sub_cell = ws['A2']
            if filter_type == '__custom__' or filter_type == 'custom':
                sub_cell.value = f"Period: {start.isoformat()} to {end.isoformat()}  |  Department: {dept}  |  Generated: {date.today().isoformat()}"
            else:
                sub_cell.value = f"Month: {month}  |  Department: {dept}  |  Generated: {date.today().isoformat()}"
            sub_cell.font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
            sub_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 20
            ws.append([]) # Row 3

            # Header Row (Row 4)
            headers = ['Emp ID', 'Full Name', 'Department', 'Date', 'Status', 'Check-In', 'Check-Out', 'Time Worked', 'Location', 'PO Number', 'PO Worked Days', 'Reporting Person', 'Zone', 'Notes']
            ws.append(headers)
            ws.row_dimensions[4].height = 24

            h_font = Font(name="Segoe UI", size=10, bold=True, color="00D4AA")
            h_fill = PatternFill(start_color="0D1F35", end_color="0D1F35", fill_type="solid")
            h_align = Alignment(horizontal="left", vertical="center")

            for col_idx in range(1, 15):
                cell = ws.cell(row=4, column=col_idx)
                cell.font = h_font
                cell.fill = h_fill
                cell.alignment = h_align

            # Borders & Fonts
            border_side = Side(style='thin', color='1F2D45')
            thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

            f_reg = Font(name="Segoe UI", size=9, color="E2E8F0")
            f_present = Font(name="Segoe UI", size=9, bold=True, color="10B981")
            f_location = Font(name="Segoe UI", size=9, bold=True, color="00D4AA")

            fill_dark = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
            fill_darker = PatternFill(start_color="1A2235", end_color="1A2235", fill_type="solid")

            # Append Data
            for idx, r in enumerate(rows):
                row_data = [
                    str(r.employee_id or ''),
                    str(r.full_name or ''),
                    str(r.department or ''),
                    str(r.date) if r.date else '',
                    str(r.status or ''),
                    str(r.check_in or '—'),
                    str(r.check_out or '—'),
                    str(r.time_worked or '—'),
                    str(r.location or '—'),
                    str(r.po_number or '—'),
                    str(r.po_worked_days if r.po_worked_days is not None else '—'),
                    str(r.reporting_person or '—'),
                    str(r.zone or '—'),
                    str(r.notes or '')
                ]
                ws.append(row_data)
                curr_row = 5 + idx
                ws.row_dimensions[curr_row].height = 22
                curr_fill = fill_darker if idx % 2 == 1 else fill_dark

                for col_idx in range(1, 15):
                    cell = ws.cell(row=curr_row, column=col_idx)
                    cell.font = f_reg
                    cell.fill = curr_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

                    if col_idx == 5:
                        cell.font = f_present
                    elif col_idx == 9:
                        cell.font = f_location

            # Auto column width
            from openpyxl.utils import get_column_letter
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in [1, 2]:
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            filename_suffix = f"{start.isoformat()}_to_{end.isoformat()}" if filter_type == 'custom' else month
            return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f'attendance_{filename_suffix}_{dept}.xlsx')
        except Exception as e:
            import traceback
            return f"<h3>Excel Export Error</h3><pre>{traceback.format_exc()}</pre>", 500

    # ── Build PDF ─────────────────────────────────────────────────────────────
    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=landscape(A4),
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Title
    title_style = ParagraphStyle('title', parent=styles['Normal'],
                                 fontSize=20, fontName='Helvetica-Bold',
                                 textColor=colors.HexColor('#00d4aa'),
                                 alignment=TA_CENTER, spaceAfter=4)
    sub_style   = ParagraphStyle('sub', parent=styles['Normal'],
                                 fontSize=11, fontName='Helvetica',
                                 textColor=colors.HexColor('#64748b'),
                                 alignment=TA_CENTER, spaceAfter=16)
    story.append(Paragraph('JD TECH \u2014 Attendance Report', title_style))
    if filter_type == 'custom':
        subtitle_text = f'Period: {start.isoformat()} to {end.isoformat()}  |  Department: {dept}  |  Generated: {date.today().isoformat()}'
    else:
        subtitle_text = f'Month: {month}  |  Department: {dept}  |  Generated: {date.today().isoformat()}'
    story.append(Paragraph(subtitle_text, sub_style))
    story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1f2d45'), spaceAfter=16))

    # Table header
    header = ['Emp ID', 'Full Name', 'Department', 'Date', 'Status', 'Check-In', 'Check-Out', 'Time Worked', 'Location', 'PO Number', 'PO Worked Days', 'Reporting Person', 'Zone', 'Notes']
    data   = [header]
    for r in rows:
        data.append([
            str(r.employee_id or ''),
            str(r.full_name or ''),
            str(r.department or ''),
            str(r.date) if r.date else '',
            str(r.status or ''),
            str(r.check_in or '—'),
            str(r.check_out or '—'),
            str(r.time_worked or '—'),
            str(r.location or '—'),
            str(r.po_number or '—'),
            str(r.po_worked_days if r.po_worked_days is not None else '—'),
            str(r.reporting_person or '—'),
            str(r.zone or '—'),
            str(r.notes or ''),
        ])

    col_widths = [1.4*cm, 2.4*cm, 2.0*cm, 1.6*cm, 1.3*cm, 1.3*cm, 1.3*cm, 1.5*cm, 2.3*cm, 2.0*cm, 1.6*cm, 2.2*cm, 1.8*cm, 2.8*cm]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        # Header row
        ('BACKGROUND',   (0,0), (-1,0), colors.HexColor('#0d1f35')),
        ('TEXTCOLOR',    (0,0), (-1,0), colors.HexColor('#00d4aa')),
        ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,0), 9),
        ('TOPPADDING',   (0,0), (-1,0), 10),
        ('BOTTOMPADDING',(0,0), (-1,0), 10),
        # Data rows
        ('BACKGROUND',   (0,1), (-1,-1), colors.HexColor('#111827')),
        ('TEXTCOLOR',    (0,1), (-1,-1), colors.HexColor('#e2e8f0')),
        ('FONTNAME',     (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',     (0,1), (-1,-1), 8),
        ('TOPPADDING',   (0,1), (-1,-1), 8),
        ('BOTTOMPADDING',(0,1), (-1,-1), 8),
        # Alternating rows
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#111827'), colors.HexColor('#1a2235')]),
        # Grid
        ('GRID',         (0,0), (-1,-1), 0.5, colors.HexColor('#1f2d45')),
        ('ALIGN',        (0,0), (-1,-1), 'LEFT'),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        # Colour status column (col 4) green, location column (col 6) teal
        ('TEXTCOLOR',    (4,1), (4,-1), colors.HexColor('#10b981')),
        ('TEXTCOLOR',    (6,1), (6,-1), colors.HexColor('#00d4aa')),
    ]))
    story.append(tbl)

    # Footer
    story.append(Spacer(1, 16))
    story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1f2d45'), spaceAfter=8))
    story.append(Paragraph(f'Total records: {len(rows)}  |  JD TECH Attendance System  |  Confidential',
                           ParagraphStyle('footer', parent=styles['Normal'],
                                          fontSize=8, fontName='Helvetica',
                                          textColor=colors.HexColor('#64748b'),
                                          alignment=TA_CENTER)))

    doc.build(story)
    buf.seek(0)
    filename_suffix = f"{start.isoformat()}_to_{end.isoformat()}" if filter_type == 'custom' else month
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'attendance_{filename_suffix}_{dept}.pdf')


@app.route('/api/locations/add', methods=['POST'])
@login_required
def add_location_api():
    data = request.get_json()
    if not data or 'name' not in data:
        return jsonify({'error': 'Invalid data'}), 400
    
    site_name = data['name'].strip()
    if not site_name:
        return jsonify({'error': 'Location name cannot be empty'}), 400
        
    existing = Location.query.filter_by(customer_name=site_name).first()
        
    if existing:
        display_name = existing.customer_name
        return jsonify({'success': True, 'name': display_name, 'message': 'Location already exists'})
        
    try:
        max_id = db.session.query(func.max(Location.customer_id)).scalar() or 0
        creator = session.get('full_name') or session.get('username') or session.get('employee_id') or 'User'
        new_loc = Location(customer_id=max_id + 1, customer_name=site_name, added_by=creator, created_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        db.session.add(new_loc)
        db.session.commit()
        return jsonify({'success': True, 'name': site_name})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ── GPS Location Verification ──────────────────────────────────────────────────

import math

def haversine_distance(lat1, lon1, lat2, lon2):
    """Return distance in metres between two GPS coordinates."""
    R = 6_371_000  # Earth radius in metres
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


@app.route('/api/verify-location', methods=['POST'])
@login_required
def verify_location():
    """Check if the given GPS coords fall within approved location radius."""
    data = request.get_json()
    if not data or 'latitude' not in data or 'longitude' not in data:
        return jsonify({'allowed': False, 'error': 'Missing coordinates'}), 400

    try:
        emp_lat = float(data['latitude'])
        emp_lon = float(data['longitude'])
    except (ValueError, TypeError):
        return jsonify({'allowed': False, 'error': 'Invalid coordinates'}), 400

    loc_name = data.get('location')
    if loc_name:
        loc = None
        for l in Location.query.all():
            display_name = l.customer_name
            if display_name == loc_name:
                loc = l
                break
        if not loc:
            return jsonify({'allowed': False, 'error': 'Location not found'}), 404
        if not (loc.latitude and loc.longitude):
            return jsonify({'allowed': True, 'bypass': True})
            
        dist = haversine_distance(emp_lat, emp_lon, loc.latitude, loc.longitude)
        if dist <= 200:
            return jsonify({
                'allowed':  True,
                'branch':   loc_name,
                'distance': round(dist),
                'radius':   200,
            })
        else:
            return jsonify({
                'allowed':        False,
                'closest_branch': loc_name,
                'distance':       round(dist),
                'radius':         200,
                'message':        f'You are {round(dist)} m away from {loc_name}. Allowed: 200 m.'
            })

    # Find unique GPS-enabled branches
    branches = []
    seen_keys = set()
    for l in Location.query.all():
        if l.latitude and l.longitude:
            key = l.customer_name
            if key not in seen_keys:
                seen_keys.add(key)
                branches.append(l)

    if not branches:
        return jsonify({'allowed': True, 'branch': None, 'message': 'No GPS-enabled locations configured; attendance allowed.'})

    closest_branch = None
    closest_dist   = float('inf')
    for branch in branches:
        dist = haversine_distance(emp_lat, emp_lon, branch.latitude, branch.longitude)
        if dist < closest_dist:
            closest_dist   = dist
            closest_branch = branch
        if dist <= 200:
            display_name = branch.customer_name
            return jsonify({
                'allowed':  True,
                'branch':   display_name,
                'distance': round(dist),
                'radius':   200,
            })

    closest_name = closest_branch.customer_name if closest_branch else "—"
    return jsonify({
        'allowed':        False,
        'closest_branch': closest_name,
        'distance':       round(closest_dist),
        'radius':         200,
        'message':        f'You are {round(closest_dist)} m away from the nearest branch ({closest_name}).'
    })


# ── Admin: Office Branch Management ───────────────────────────────────────────

@app.route('/admin/office-locations', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_office_locations():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add':
            try:
                name        = request.form['name'].strip()
                zone        = request.form.get('zone', '').strip()
                lat         = float(request.form['latitude']) if request.form.get('latitude') else None
                lon         = float(request.form['longitude']) if request.form.get('longitude') else None
                
                max_id = db.session.query(func.max(Location.customer_id)).scalar() or 0
                creator = session.get('full_name') or session.get('username') or 'Admin'
                
                db.session.add(Location(
                    customer_id=max_id + 1,
                    customer_name=name, zone=zone,
                    latitude=lat, longitude=lon,
                    created_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    added_by=creator
                ))
                db.session.commit()
                flash(f'Location "{name}" added successfully.', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error adding location: {e}', 'error')

        elif action == 'delete':
            loc_id = request.form.get('location_id')
            loc = Location.query.get(loc_id)
            if loc:
                name = loc.customer_name
                # Delete all POs for this customer name
                PO.query.filter_by(customer_name=name).delete()
                # Delete location
                db.session.delete(loc)
                db.session.commit()
                flash(f'Location "{name}" deleted.', 'success')

        elif action == 'toggle':
            loc_id = request.form.get('location_id')
            loc = Location.query.get(loc_id)
            if loc:
                if loc.latitude or loc.longitude:
                    # Clear coords to disable GPS
                    loc.latitude = None
                    loc.longitude = None
                    db.session.commit()
                    flash(f'GPS verification disabled for "{loc.customer_name}".', 'success')
                else:
                    flash(f'To enable GPS verification for "{loc.customer_name}", please edit the location and enter Latitude and Longitude.', 'warning')

        elif action == 'edit':
            loc_id = request.form.get('location_id')
            loc = Location.query.get(loc_id)
            if loc:
                try:
                    old_name = loc.customer_name
                    new_name = request.form['name'].strip()
                    zone     = request.form.get('zone', '').strip()
                    lat      = float(request.form['latitude']) if request.form.get('latitude') else None
                    lon      = float(request.form['longitude']) if request.form.get('longitude') else None
                    
                    loc.customer_name = new_name
                    loc.zone          = zone
                    loc.latitude      = lat
                    loc.longitude     = lon
                    
                    if old_name != new_name:
                        # Cascade update to PO table
                        PO.query.filter_by(customer_name=old_name).update({PO.customer_name: new_name})

                    db.session.commit()
                    flash(f'Location "{new_name}" updated.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error updating location: {e}', 'error')

        elif action == 'add_po':
            try:
                loc_id = request.form.get('location_id')
                zone   = request.form.get('zone', '').strip()
                po_num = request.form['po_number'].strip()
                days   = request.form.get('days_needed', '').strip()
                rp     = request.form.get('reporting_person', '').strip()
                
                base_loc = Location.query.get(loc_id)
                if base_loc and po_num and zone:
                    max_sno = db.session.query(func.max(PO.s_no)).scalar() or 0
                    db.session.add(PO(
                        s_no=max_sno + 1,
                        customer_name=base_loc.customer_name,
                        zone=zone,
                        po_number=po_num,
                        days=days,
                        reporting_person=rp
                    ))
                    db.session.commit()
                    flash(f'PO "{po_num}" added successfully.', 'success')
                else:
                    flash('PO Number and Zone cannot be empty.', 'error')
            except Exception as e:
                db.session.rollback()
                flash(f'Error adding PO: {e}', 'error')

        elif action == 'delete_po':
            po_id = request.form.get('po_id')
            po_row = PO.query.get(po_id)
            if po_row:
                number = po_row.po_number
                db.session.delete(po_row)
                db.session.commit()
                flash(f'PO "{number}" deleted.', 'success')

        return redirect(url_for('admin_office_locations'))

    raw_locations = Location.query.order_by(Location.customer_name).all()
    branches = []
    for loc in raw_locations:
        pos_list = PO.query.filter_by(customer_name=loc.customer_name).all()
        branches.append({
            'id': loc.id,
            'customer_id': loc.customer_id,
            'customer_name': loc.customer_name,
            'name': loc.customer_name,
            'zone': loc.zone,
            'latitude': loc.latitude,
            'longitude': loc.longitude,
            'added_by': loc.added_by,
            'gps_enabled': bool(loc.latitude and loc.longitude),
            'pos': [{
                'id': p.id,
                'zone': p.zone,
                'number': p.po_number,
                'days_needed': p.days,
                'reporting_person': p.reporting_person,
                'days_worked': Attendance.query.filter(Attendance.po_number == p.po_number, Attendance.status.in_(['Present', 'Half Day'])).count()
            } for p in pos_list],
            'reporting_persons': []
        })
    return render_template('admin_office_locations.html', branches=branches)

@app.route('/sync-from-sheets')
@login_required
def sync_from_sheets_route():
    try:
        pull_from_sheets(app)
        flash('🔄 Successfully synchronized latest data & deletions from Google Sheets!', 'success')
    except Exception as e:
        flash(f'Error syncing from Google Sheets: {e}', 'error')
    return redirect(request.referrer or url_for('my_attendance'))

if __name__ == '__main__':
    import socket
    with app.app_context():
        init_db()

    # Detect local network IP automatically
    try:
        hostname = socket.gethostname()
        local_ip = socket.gethostbyname(hostname)
    except Exception:
        local_ip = '0.0.0.0'

    import os
    port = int(os.environ.get('PORT', 5001))
    print("=" * 60)
    print("  JD TECH Attendance System — STARTING")
    print("=" * 60)
    print(f"  Local Access  : http://localhost:{port}")
    print(f"  Network Link  : http://{local_ip}:{port}")
    print()
    print("  Share the Network Link with your team members.")
    print("  Make sure all devices are on the same WiFi/LAN.")
    print("=" * 60)
    serve(app, host='0.0.0.0', port=port, threads=4)
